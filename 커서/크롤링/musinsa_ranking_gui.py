"""
무신사 랭킹 크롤링 GUI
- [시작] 버튼: 실시간 크롤링 후 엑셀 데이터 준비
- [저장] 버튼: 엑셀 파일로 저장
"""

import sys
from pathlib import Path

# 스크립트 폴더를 경로에 추가 (다른 경로에서 실행해도 모듈 인식)
sys.path.insert(0, str(Path(__file__).resolve().parent))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from musinsa_ranking_crawler import crawl_musinsa_ranking, DEFAULT_URL, MAX_ITEMS
except ImportError:
    DEFAULT_URL = ""
    MAX_ITEMS = 100
    crawl_musinsa_ranking = None


def items_to_excel(items: list[dict], filepath: str) -> None:
    """크롤링 결과를 엑셀 파일로 저장"""
    if not HAS_EXCEL:
        raise ImportError(
            "openpyxl이 설치되지 않았습니다.\n"
            "명령 프롬프트에서 실행: pip install openpyxl"
        )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "무신사 랭킹"

    headers = ["순위", "브랜드", "상품명", "가격", "상품링크"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.get("rank", row_idx - 1))
        ws.cell(row=row_idx, column=2, value=item.get("brand", ""))
        ws.cell(row=row_idx, column=3, value=item.get("product", ""))
        ws.cell(row=row_idx, column=4, value=item.get("price", ""))
        ws.cell(row=row_idx, column=5, value=item.get("link", ""))

    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(
            12, min(50, max(len(str(h)) for h in headers) + 2)
        )
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["E"].width = 60

    wb.save(filepath)


class MusinsaRankingApp:
    def __init__(self):
        if crawl_musinsa_ranking is None:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "모듈 오류",
                "musinsa_ranking_crawler 모듈을 찾을 수 없습니다.\n"
                "같은 폴더에 musinsa_ranking_crawler.py 파일이 있는지 확인하세요.",
            )
            sys.exit(1)

        if not HAS_EXCEL:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "패키지 필요",
                "openpyxl이 설치되지 않았습니다.\n\n"
                "명령 프롬프트에서 실행:\n  pip install openpyxl",
            )
            sys.exit(1)

        self.root = tk.Tk()
        self.root.title("무신사 랭킹 크롤링")
        self.root.geometry("580x420")
        self.root.minsize(500, 360)

        self.items: list[dict] = []
        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 14, "pady": 8}

        # 상단 안내
        info_frame = ttk.LabelFrame(self.root, text="무신사 랭킹", padding=10)
        info_frame.pack(fill=tk.X, **pad)

        ttk.Label(
            info_frame,
            text="랭킹 페이지를 크롤링한 뒤 엑셀로 저장할 수 있습니다. (최대 100건)",
            foreground="gray",
        ).pack(anchor=tk.W)

        # 버튼 영역
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(fill=tk.X, **pad)

        self.btn_start = ttk.Button(
            btn_frame,
            text="시작 (크롤링)",
            command=self._on_start,
        )
        self.btn_start.pack(side=tk.LEFT, padx=(0, 8))

        self.btn_save = ttk.Button(
            btn_frame,
            text="저장 (엑셀)",
            command=self._on_save,
            state=tk.DISABLED,
        )
        self.btn_save.pack(side=tk.LEFT)

        # 결과 테이블
        table_frame = ttk.LabelFrame(self.root, text="크롤링 결과 미리보기", padding=8)
        table_frame.pack(fill=tk.BOTH, expand=True, **pad)

        columns = ("rank", "brand", "product", "price")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="headings",
            height=10,
            selectmode="browse",
        )
        self.tree.heading("rank", text="순위")
        self.tree.heading("brand", text="브랜드")
        self.tree.heading("product", text="상품명")
        self.tree.heading("price", text="가격")

        self.tree.column("rank", width=50)
        self.tree.column("brand", width=120)
        self.tree.column("product", width=280)
        self.tree.column("price", width=80)

        scroll_y = ttk.Scrollbar(table_frame)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scroll_y.set)
        scroll_y.configure(command=self.tree.yview)

        self.tree.pack(fill=tk.BOTH, expand=True)

        # 상태 표시
        self.status_var = tk.StringVar()
        self.status_var.set("시작 버튼을 눌러 크롤링을 실행하세요.")
        status_label = ttk.Label(
            self.root,
            textvariable=self.status_var,
            foreground="gray",
            wraplength=520,
        )
        status_label.pack(fill=tk.X, **pad)

    def _update_tree(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for item in self.items:
            self.tree.insert("", tk.END, values=(
                item.get("rank", ""),
                item.get("brand", ""),
                (item.get("product", "") or "")[:60] + ("..." if len(item.get("product", "") or "") > 60 else ""),
                item.get("price", ""),
            ))

    def _on_start(self):
        self.status_var.set("크롤링 중... (첫 시도 실패 시 브라우저 자동 실행, 20초 정도 소요될 수 있음)")
        self.btn_start.configure(state=tk.DISABLED)
        self.root.update()

        def do_crawl():
            try:
                items = crawl_musinsa_ranking(max_items=MAX_ITEMS)
                if not items:
                    items = crawl_musinsa_ranking(use_backup=True, max_items=MAX_ITEMS)
                return items, None
            except RuntimeError as e:
                return [], e
            except Exception as e:
                return [], RuntimeError(f"크롤링 중 오류 발생: {e}")

        items, err = do_crawl()

        self.btn_start.configure(state=tk.NORMAL)
        if err:
            err_msg = str(err)
            self.status_var.set(f"오류: {err_msg}")
            messagebox.showerror("크롤링 오류", err_msg)
            return

        self.items = items
        self._update_tree()

        if items:
            self.btn_save.configure(state=tk.NORMAL)
            self.status_var.set(f"크롤링 완료: {len(items)}개 상품 (저장 버튼으로 엑셀 저장)")
        else:
            self.status_var.set("수집된 상품이 없습니다. 무신사 페이지 구조가 변경되었을 수 있습니다.")
            messagebox.showwarning("알림", "수집된 상품이 없습니다.\n페이지 구조가 변경되었거나 접근이 제한되었을 수 있습니다.")

    def _on_save(self):
        if not self.items:
            messagebox.showwarning("알림", "먼저 시작 버튼으로 크롤링을 실행하세요.")
            return

        from datetime import datetime
        default_name = f"무신사_랭킹_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="엑셀 파일로 저장",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("엑셀 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if not save_path:
            return

        try:
            items_to_excel(self.items, save_path)
            self.status_var.set(f"저장 완료: {save_path}")
            messagebox.showinfo("저장 완료", f"엑셀 파일로 저장했습니다.\n{save_path}")
        except Exception as e:
            self.status_var.set(f"저장 오류: {e}")
            messagebox.showerror("저장 오류", str(e))

    def run(self):
        self.root.mainloop()


def main():
    app = MusinsaRankingApp()
    app.run()


if __name__ == "__main__":
    main()
